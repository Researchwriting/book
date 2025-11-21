"""
Research Instrument Designer
Generates questionnaires/interview guides based on objectives and methodology
"""
from .llm import LLMClient
import os

class InstrumentDesigner:
    def __init__(self, llm_client):
        self.llm = llm_client
    
    def design_instrument(self, topic, case_study, objectives, research_questions, methodology_type="quantitative"):
        """
        Design a data collection instrument based on research objectives and questions.
        
        Args:
            topic: Research topic
            case_study: Case study location
            objectives: Research objectives from Chapter 1
            research_questions: Research questions from Chapter 1
            methodology_type: "quantitative", "qualitative", or "mixed"
        
        Returns:
            Instrument content as markdown
        """
        print("\n📋 Designing data collection instrument...")
        
        if methodology_type.lower() == "quantitative":
            instrument = self._design_questionnaire(topic, case_study, objectives, research_questions)
        elif methodology_type.lower() == "qualitative":
            instrument = self._design_interview_guide(topic, case_study, objectives, research_questions)
        else:  # mixed methods
            questionnaire = self._design_questionnaire(topic, case_study, objectives, research_questions)
            interview_guide = self._design_interview_guide(topic, case_study, objectives, research_questions)
            instrument = f"{questionnaire}\n\n---\n\n{interview_guide}"
        
        return instrument
    
    def _design_questionnaire(self, topic, case_study, objectives, research_questions):
        """Design a structured questionnaire"""
        print("  📝 Designing questionnaire...")
        
        prompt = f"""
You are designing a research questionnaire for a PhD study.

TOPIC: {topic}
CASE STUDY: {case_study}

RESEARCH OBJECTIVES:
{objectives}

RESEARCH QUESTIONS:
{research_questions}

TASK: Design a comprehensive questionnaire that will collect data to address ALL the research objectives and answer ALL the research questions.

STRUCTURE:
1. **Cover Page**: Title, researcher info, purpose statement
2. **Instructions**: How to complete the questionnaire
3. **Section A: Demographic Information**
   - Age, gender, education, occupation, etc.
   - 5-7 questions

4. **Section B-E**: One section per research objective/question
   - Each section should have 8-12 questions
   - Use a mix of question types:
     * Likert scale (1-5: Strongly Disagree to Strongly Agree)
     * Multiple choice
     * Yes/No
     * Open-ended (for qualitative insights)
   
5. **Closing**: Thank you message

REQUIREMENTS:
- Questions must be clear, unbiased, and directly linked to objectives
- Use appropriate scales (Likert, rating, etc.)
- Include reverse-coded items to check consistency
- Ensure questions can generate quantitative data for analysis
- Add a note after each section showing which objective/question it addresses

FORMAT: Use markdown with clear headings and formatting.

LENGTH: Comprehensive questionnaire (50-80 questions total)
"""
        
        questionnaire = self.llm.generate(
            prompt,
            system_prompt="You are an expert research methodologist designing data collection instruments.",
            max_tokens=4096
        )
        
        return f"# APPENDIX A: RESEARCH QUESTIONNAIRE\n\n{questionnaire}"
    
    def _design_interview_guide(self, topic, case_study, objectives, research_questions):
        """Design a semi-structured interview guide"""
        print("  🎤 Designing interview guide...")
        
        prompt = f"""
You are designing a semi-structured interview guide for a PhD study.

TOPIC: {topic}
CASE STUDY: {case_study}

RESEARCH OBJECTIVES:
{objectives}

RESEARCH QUESTIONS:
{research_questions}

TASK: Design a comprehensive interview guide that will collect qualitative data to address ALL the research objectives and answer ALL the research questions.

STRUCTURE:
1. **Introduction Script**: How to introduce the interview
2. **Informed Consent**: Brief consent statement
3. **Warm-up Questions**: 2-3 easy questions to build rapport
4. **Main Sections**: One section per research objective
   - Each section: 5-8 main questions
   - Include 2-3 probing questions per main question
   - Use open-ended questions

5. **Closing Questions**: Final thoughts, anything to add
6. **Thank You Script**

REQUIREMENTS:
- Questions must be open-ended and exploratory
- Include probing questions (e.g., "Can you elaborate?", "Why do you think that?")
- Questions should encourage detailed narratives
- Link each section to specific objectives
- Include prompts for the interviewer

FORMAT: Use markdown with clear headings.

LENGTH: Comprehensive guide (25-40 main questions with probes)
"""
        
        interview_guide = self.llm.generate(
            prompt,
            system_prompt="You are an expert qualitative researcher designing interview protocols.",
            max_tokens=4096
        )
        
        return f"# APPENDIX B: INTERVIEW GUIDE\n\n{interview_guide}"
    
    def save_instrument(self, instrument, output_dir="thesis/appendices"):
        """Save the instrument to appendices folder"""
        os.makedirs(output_dir, exist_ok=True)
        filename = f"{output_dir}/Research_Instrument.md"
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(instrument)
        
        print(f"  ✅ Instrument saved: {filename}")
        return filename
    
    def generate_simulated_data(self, instrument, sample_size, topic, case_study):
        """
        Generate simulated/realistic data based on the instrument.
        This will be used for Chapter 4 data presentation.
        """
        print(f"\n📊 Generating simulated data (N={sample_size})...")
        
        prompt = f"""
You are generating realistic simulated data for a PhD thesis.

TOPIC: {topic}
CASE STUDY: {case_study}
SAMPLE SIZE: {sample_size}

INSTRUMENT:
{instrument[:3000]}... [truncated]

TASK: Generate realistic simulated data that would result from administering this instrument to {sample_size} respondents in {case_study}.

REQUIREMENTS:
1. **Demographic Data**: Generate realistic distributions
   - Example: Age groups, gender ratios, education levels

2. **Response Data**: For each question/variable:
   - Generate realistic frequency distributions
   - Ensure data makes sense for the context of {case_study}
   - Include some variability (not all perfect responses)

3. **Format as Tables**: Present data in markdown tables
   - Frequency tables for categorical data
   - Descriptive statistics for Likert scales (mean, SD)
   - Cross-tabulations where relevant

4. **Statistical Results**: Include realistic statistical findings:
   - Correlation coefficients (if applicable)
   - Regression results (if applicable)
   - Chi-square tests (if applicable)
   - T-tests or ANOVA (if applicable)

OUTPUT FORMAT:
```markdown
## Simulated Data Summary

### Demographics
[Table with demographic distributions]

### Variable 1: [Name]
[Frequency table or descriptive statistics]

### Variable 2: [Name]
[Frequency table or descriptive statistics]

... (continue for all major variables)

### Statistical Relationships
[Correlation matrix, regression results, etc.]
```

Make the data realistic and consistent with the research context.
"""
        
        simulated_data = self.llm.generate(
            prompt,
            system_prompt="You are a statistician generating realistic research data.",
            max_tokens=4096
        )
        
        return simulated_data
    
    def save_simulated_data(self, data, output_dir="thesis/data"):
        """Save simulated data for use in Chapter 4"""
        os.makedirs(output_dir, exist_ok=True)
        filename = f"{output_dir}/Simulated_Data.md"
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("# SIMULATED RESEARCH DATA\n\n")
            f.write("**Note**: This is simulated data for thesis demonstration purposes.\n\n")
            f.write(data)
        
        print(f"  ✅ Simulated data saved: {filename}")
        return filename
    
    def generate_dataset(self, instrument, sample_size, topic, case_study, output_dir="thesis/data"):
        """
        Generate actual CSV/Excel dataset based on the instrument.
        Works for quantitative, qualitative, or mixed methods.
        """
        print(f"\n📊 Generating dataset files (N={sample_size})...")
        
        os.makedirs(output_dir, exist_ok=True)
        
        # Determine methodology type from instrument
        if "QUESTIONNAIRE" in instrument.upper():
            dataset = self._generate_quantitative_dataset(instrument, sample_size, topic, case_study)
            csv_file = f"{output_dir}/Quantitative_Data.csv"
            excel_file = f"{output_dir}/Quantitative_Data.xlsx"
        elif "INTERVIEW" in instrument.upper():
            dataset = self._generate_qualitative_dataset(instrument, sample_size, topic, case_study)
            csv_file = f"{output_dir}/Qualitative_Data.csv"
            excel_file = f"{output_dir}/Qualitative_Data.xlsx"
        else:
            # Mixed methods - generate both
            quant_data = self._generate_quantitative_dataset(instrument, sample_size, topic, case_study)
            qual_data = self._generate_qualitative_dataset(instrument, int(sample_size * 0.3), topic, case_study)
            
            self._save_csv(quant_data, f"{output_dir}/Quantitative_Data.csv")
            self._save_excel(quant_data, f"{output_dir}/Quantitative_Data.xlsx")
            self._save_csv(qual_data, f"{output_dir}/Qualitative_Data.csv")
            self._save_excel(qual_data, f"{output_dir}/Qualitative_Data.xlsx")
            
            print(f"  ✅ Quantitative CSV: {output_dir}/Quantitative_Data.csv")
            print(f"  ✅ Quantitative Excel: {output_dir}/Quantitative_Data.xlsx")
            print(f"  ✅ Qualitative CSV: {output_dir}/Qualitative_Data.csv")
            print(f"  ✅ Qualitative Excel: {output_dir}/Qualitative_Data.xlsx")
            return
        
        # Save to CSV and Excel
        self._save_csv(dataset, csv_file)
        self._save_excel(dataset, excel_file)
        
        print(f"  ✅ CSV: {csv_file}")
        print(f"  ✅ Excel: {excel_file}")
        
        return csv_file, excel_file
    
    def _generate_quantitative_dataset(self, instrument, sample_size, topic, case_study):
        """Generate quantitative dataset (surveys/questionnaires)"""
        import random
        import csv
        from datetime import datetime, timedelta
        
        # Extract variables from instrument (simplified - in production, parse instrument)
        dataset = []
        
        # Generate header row
        headers = [
            "RespondentID", "Date", "Age", "Gender", "Education", "Occupation",
            "Q1_Employment_Status", "Q2_War_Impact", "Q3_Income_Change",
            "Q4_Job_Security", "Q5_Economic_Stability", "Q6_Future_Outlook",
            "Q7_Government_Support", "Q8_Skills_Training", "Q9_Migration",
            "Q10_Family_Impact"
        ]
        
        # Add Likert scale questions (1-5)
        for i in range(11, 31):
            headers.append(f"Q{i}_Likert")
        
        dataset.append(headers)
        
        # Generate data rows
        for i in range(1, sample_size + 1):
            row = [
                f"R{i:04d}",  # RespondentID
                (datetime.now() - timedelta(days=random.randint(1, 90))).strftime("%Y-%m-%d"),  # Date
                random.randint(18, 65),  # Age
                random.choice(["Male", "Female"]),  # Gender
                random.choice(["Primary", "Secondary", "Diploma", "Degree", "Postgraduate"]),  # Education
                random.choice(["Employed", "Unemployed", "Self-employed", "Student"]),  # Occupation
                random.choice(["Full-time", "Part-time", "Casual", "Unemployed"]),  # Q1
                random.randint(1, 5),  # Q2 (1=No impact, 5=Severe impact)
                random.choice(["Increased", "Decreased", "No change"]),  # Q3
                random.randint(1, 5),  # Q4
                random.randint(1, 5),  # Q5
                random.randint(1, 5),  # Q6
                random.randint(1, 5),  # Q7
                random.randint(1, 5),  # Q8
                random.choice(["Yes", "No", "Considering"]),  # Q9
                random.randint(1, 5),  # Q10
            ]
            
            # Add Likert responses (Q11-Q30)
            for _ in range(20):
                row.append(random.randint(1, 5))
            
            dataset.append(row)
        
        return dataset
    
    def _generate_qualitative_dataset(self, instrument, sample_size, topic, case_study):
        """Generate qualitative dataset (interviews/focus groups)"""
        import random
        from datetime import datetime, timedelta
        
        dataset = []
        
        # Header row
        headers = [
            "ParticipantID", "Date", "Age", "Gender", "Occupation",
            "Interview_Duration_Minutes", "Location", "Theme1_Response",
            "Theme2_Response", "Theme3_Response", "Key_Quotes", "Notes"
        ]
        
        dataset.append(headers)
        
        # Sample responses for qualitative data
        theme1_responses = [
            "Expressed significant concern about job security due to ongoing conflict",
            "Mentioned loss of employment opportunities in the region",
            "Described challenges in finding stable work",
            "Highlighted impact of war on local businesses"
        ]
        
        theme2_responses = [
            "Discussed migration to urban areas for better opportunities",
            "Mentioned family separation due to employment search",
            "Described reliance on informal economy",
            "Highlighted need for skills training programs"
        ]
        
        theme3_responses = [
            "Emphasized need for government intervention",
            "Discussed community support mechanisms",
            "Mentioned international aid dependency",
            "Highlighted youth unemployment crisis"
        ]
        
        quotes = [
            "The war has destroyed everything we built",
            "I lost my job when the conflict started",
            "We need peace before we can think about employment",
            "Young people have no future here"
        ]
        
        # Generate data rows
        for i in range(1, sample_size + 1):
            row = [
                f"P{i:03d}",  # ParticipantID
                (datetime.now() - timedelta(days=random.randint(1, 60))).strftime("%Y-%m-%d"),
                random.randint(25, 60),  # Age
                random.choice(["Male", "Female"]),
                random.choice(["Teacher", "Farmer", "Trader", "Civil Servant", "Unemployed"]),
                random.randint(30, 90),  # Interview duration
                random.choice(["Juba", "Malakal", "Wau", "Bor"]),
                random.choice(theme1_responses),
                random.choice(theme2_responses),
                random.choice(theme3_responses),
                random.choice(quotes),
                f"Interview conducted in local language, translated. Participant showed emotional response."
            ]
            
            dataset.append(row)
        
        return dataset
    
    def _save_csv(self, dataset, filename):
        """Save dataset to CSV file"""
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(dataset)
    
    def _save_excel(self, dataset, filename):
        """Save dataset to Excel file"""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Research Data"
            
            # Add data
            for row_idx, row in enumerate(dataset, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Style header row
                    if row_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except ImportError:
            print("  ⚠️  openpyxl not installed. Skipping Excel file generation.")
            print("     Install with: pip install openpyxl")
